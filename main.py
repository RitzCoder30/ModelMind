from flask import Flask, render_template, request, session
import pandas as pd
from transformers import pipeline
import io, os
from openpyxl import load_workbook
from difflib import get_close_matches
import logging
from uuid import uuid4
import spacy

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "default_secret_key")
app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024  # 5MB max upload

pipe = pipeline("text2text-generation", model="google/flan-t5-small")

# Define common Excel error strings that should be ignored during data extraction
errors = {"#DIV/0!", "#VALUE!", "#REF!", "#NAME?", "#NUM!", "#N/A", "#NULL!"}

# This allows the app to retain the DataFrame in memory for subsequent chat queries
uploaded_files = {}

# Configure logging for better visibility into application flow and potential issues
logging.basicConfig(level=logging.INFO)

# Load the spaCy English NLP model for natural language understanding.
# This model is used to parse user questions and identify key entities like metrics, departments, and years.
nlp = spacy.load("en_core_web_sm")

def normalize_col(s):
    return (
        str(s)
        .strip()
        .lower()
        .replace("&", "and")
        .replace(".", "")
        .replace(" ", "_")
        .replace("-", "_")
    )


def dataframe_to_text_table(df):
    # Create a copy, fill any NaN values with empty strings, and convert all data to string type
    df = df.copy().fillna("").astype(str)
    headers = list(df.columns)
    # Create the header row and a separator row for markdown table formatting
    lines = [" | ".join(headers), " | ".join(["---"] * len(headers))]
    # Add each row's data to the table
    for _, row in df.iterrows():
        lines.append(" | ".join(row.values))
    return "\n".join(lines)

def fuzzy_match(label, aliases, cutoff=0.8):
    return any(get_close_matches(label.lower(), [alias.lower() for alias in aliases], n=1, cutoff=cutoff))

def extract_financial_summary(df):
    df = df.copy()
    df.columns = [normalize_col(col) for col in df.columns]

    label_col = df.columns[
        0]  # Assumes the first column contains the financial metric labels (e.g., "Category_Revenue")

    # Attempt to find a column explicitly named 'total' or containing 'total'
    total_col = next((col for col in df.columns if 'total' in col), None)

    # Identify other columns that might represent departmental or year-based breakdowns
    # These are columns that are not the label column and not the identified total column
    dept_cols = [col for col in df.columns if col not in {label_col, total_col}]

    # mapping of financial metrics to their common aliases
    row_label_map = {
        "Revenue": ["Revenue", "Total Revenue", "Revenues", "Sales", "Category_Revenue"],
        "Cost of Goods Sold": ["COGS", "Cost of Goods Sold"],
        "Gross Profit": ["Gross Profit", "Gross Margin"],
        "Operating Expenses": ["Operating Expenses", "OPEX", "Operating Costs"],
        "Operating Profit": ["Operating Profit", "Operating Income", "EBIT"],
        "Net Income": ["Net Income", "Net Profit"]
    }

    extracted = {}  # Dictionary to store extracted financial figures

    # Iterate through each row of the DataFrame to find and extract metric values
    for _, row in df.iterrows():
        label = str(row[label_col]).strip()
        # Attempt to match the row label to one of the canonical financial metrics
        for canonical, aliases in row_label_map.items():
            if fuzzy_match(label, aliases):
                total_val = None
                if total_col and total_col in row and pd.notnull(row[total_col]) and row[total_col] not in errors:
                    try:
                        total_val = float(row[total_col])
                    except ValueError:
                        logging.info(f"Skipping non-numeric total value for {label}: {row[total_col]}")

                if total_val is not None:
                    depts = []
                    # Collect breakdown values from department/year columns
                    for dept_col in dept_cols:
                        val = row.get(dept_col)
                        if pd.notnull(val) and val not in errors:
                            try:
                                val = float(val)
                                # Format breakdown for summary, e.g., "Sales_2023: $100.00"
                                depts.append(f"{dept_col.replace('_', ' ').title()}: ${val:,.2f}")
                            except ValueError:
                                logging.info(f"Skipping non-numeric value for {dept_col} under {label}: {val}")
                                continue
                    extracted[canonical] = (total_val, depts)
                break  # Move to the next row once a metric match is found

    if not extracted:
        return "Could not extract key financial figures."

    # Format the extracted figures into a human-readable summary string
    lines = []
    for k, (total, dept_vals) in extracted.items():
        if dept_vals:
            lines.append(f"- {k}: ${total:,.2f} ({'; '.join(dept_vals)})")
        else:
            lines.append(f"- {k}: ${total:,.2f}")
    return "Key Financial Figures:\n" + "\n".join(lines)


def extract_detailed_summaries(df):
    df_processed = df.copy()
    df_processed.columns = [normalize_col(col) for col in df.columns]
    logging.info(f"Columns after normalization in extract_detailed_summaries: {df.columns.tolist()}")

    label_col = df_processed.columns[0]

    # mapping for financial metrics
    metric_aliases = {
        "revenue": ["revenue", "total_revenue", "revenues", "sales"],
        "cost_of_goods_sold": ["cogs", "cost of goods sold", "cost_of_goods_sold"],
        "gross_profit": ["gross profit", "gross_margin"],
        "operating_expenses": ["operating expenses", "opex", "operating_costs"],
        "operating_profit": ["operating profit", "operating income", "ebit"],
        "net_income": ["net income", "net profit"]
    }

    detailed_summary = {}  # The main dictionary to store the structured data

    # Revenue data is in column headers.
    detailed_summary['revenue'] = {}

    for col_name in df_processed.columns:
        if 'total' in col_name:
            try:
                # Assuming 'total_XXXX.X' format, where XXXX.X is the value
                total_val_str = col_name.split('total_')[-1]
                total_val = float(total_val_str)
                detailed_summary['revenue']['total'] = total_val
                logging.info(f"Extracted total revenue from column name '{col_name}': ${total_val:,.2f}")
            except (ValueError, IndexError):
                logging.warning(f"Could not extract numeric total from column name: {col_name}")

        elif 'dept' in col_name and '_' in col_name:
            try:
                # Assuming 'dept_XXXX.X' format
                parts = col_name.rsplit('_', 1)
                if len(parts) == 2:
                    dept_name = parts[0]
                    dept_val_str = parts[1]
                    dept_val = float(dept_val_str)

                    # Store directly under department name for simplicity
                    detailed_summary['revenue'][dept_name] = dept_val
                    logging.info(f"Extracted revenue for '{dept_name}' from column name '{col_name}': ${dept_val:,.2f}")
            except (ValueError, IndexError):
                logging.warning(f"Could not extract numeric department revenue from column name: {col_name}")

    # --- General Handling for other metrics (Cost of Goods Sold, Gross Profit, etc.) from rows ---
    for _, row in df_processed.iterrows():
        row_label = str(row[label_col]).strip()
        logging.info(f"Processing row label: '{row_label}' from column '{label_col}' for general metrics.")
        matched_metric = None

        for canonical_metric, aliases in metric_aliases.items():
            if canonical_metric != "revenue" and fuzzy_match(row_label, aliases):
                matched_metric = canonical_metric
                logging.info(f"Matched '{row_label}' to canonical metric: '{canonical_metric}'")
                break

        if matched_metric:
            if matched_metric not in detailed_summary:
                detailed_summary[matched_metric] = {}

            for col_name in df_processed.columns:
                if col_name == label_col:
                    continue

                value = row.get(col_name)
                if pd.notnull(value) and value not in errors:
                    try:
                        float_value = float(value)
                    except ValueError:
                        logging.info(
                            f"Skipping non-numeric value '{value}' in column '{col_name}' for metric '{matched_metric}'.")
                        continue

                    if 'total' in col_name:
                        detailed_summary[matched_metric]['total'] = float_value
                        continue

                    parts = col_name.split('_')
                    if len(parts) >= 2:
                        parent_key = parts[0]
                        child_key = "_".join(parts[1:])

                        if parent_key not in detailed_summary[matched_metric]:
                            detailed_summary[matched_metric][parent_key] = {}
                        detailed_summary[matched_metric][parent_key][child_key] = float_value
                    else:
                        detailed_summary[matched_metric][col_name] = float_value

    logging.info(f"Detailed summary built: {detailed_summary}")
    return detailed_summary


def get_error_cells(file_bytes):
    workbook = load_workbook(filename=io.BytesIO(file_bytes), data_only=False)
    sheet = workbook.active
    return [
        cell.coordinate
        for row in sheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.strip() in errors
    ]


financial_formulas = {
    "gross_profit": {
        "formula": "Revenue - Cost of Goods Sold",
        "calculation_func": lambda data, department=None: calculate_gross_profit(data, department)
    },
    "net_income": {
        "formula": "Gross Profit - Operating Expenses",
        "calculation_func": lambda data, department=None: calculate_net_income(data, department)
    }
}


def calculate_gross_profit(detailed_summary, department=None):
    revenue_data = detailed_summary.get('revenue', {})
    cogs_data = detailed_summary.get('cost_of_goods_sold', {})

    if department:
        # Attempt to find the specific department's revenue
        dept_revenue = 0
        if department in revenue_data and isinstance(revenue_data[department], (int, float)):
            dept_revenue = revenue_data[department]
        else:  # Look for department within nested structures
            for key, val_dict in revenue_data.items():
                if isinstance(val_dict, dict) and department in val_dict:
                    dept_revenue = val_dict[department]
                    break

        # Attempt to find the specific department's COGS
        dept_cogs = 0
        if department in cogs_data and isinstance(cogs_data[department], (int, float)):
            dept_cogs = cogs_data[department]
        else:  # Look for department within nested structures
            for key, val_dict in cogs_data.items():
                if isinstance(val_dict, dict) and department in val_dict:
                    dept_cogs = val_dict[department]
                    break

        if dept_revenue is not None and dept_cogs is not None:
            return dept_revenue - dept_cogs
        return None
    else:
        # Calculate total gross profit
        total_revenue = revenue_data.get('total')
        total_cogs = cogs_data.get('total')
        if total_revenue is not None and total_cogs is not None:
            return total_revenue - total_cogs
        return None


def calculate_net_income(detailed_summary, department=None):
    gross_profit = calculate_gross_profit(detailed_summary, department)
    if gross_profit is None:
        return None

    opex_data = detailed_summary.get('operating_expenses', {})

    if department:
        dept_opex = 0
        if department in opex_data and isinstance(opex_data[department], (int, float)):
            dept_opex = opex_data[department]
        else:  # Look for department within nested structures
            for key, val_dict in opex_data.items():
                if isinstance(val_dict, dict) and department in val_dict:
                    dept_opex = val_dict[department]
                    break

        if dept_opex is not None:
            return gross_profit - dept_opex
        return None
    else:
        total_opex = opex_data.get('total')
        if total_opex is not None:
            return gross_profit - total_opex
        return None


def query_detailed_summary(question, detailed_summary):
    doc = nlp(question.lower())

    # Keywords for identifying financial metrics in the question
    metric_keywords = {
        "revenue": ["revenue", "sales", "income", "revenues"],
        "cost_of_goods_sold": ["cogs", "cost of goods sold", "cost_of_goods_sold"],
        "gross_profit": ["gross profit", "margin", "gross_margin"],
        "operating_expenses": ["operating cost", "opex", "operating expenses", "operating_costs"],
        "operating_profit": ["operating profit", "operating income", "ebit"],
        "net_income": ["net income", "net profit"]
    }

    requested_metric = None
    # Identify the main financial metric requested in the question
    for token in doc:
        for canonical, aliases in metric_keywords.items():
            if fuzzy_match(token.text, aliases, cutoff=0.85):
                requested_metric = canonical
                break
        if requested_metric:
            break

    if not requested_metric:
        for chunk in doc.noun_chunks:
            for canonical, aliases in metric_keywords.items():
                if fuzzy_match(chunk.text, aliases, cutoff=0.85):
                    requested_metric = canonical
                    break
            if requested_metric:
                break

    logging.info(f"NLP detected requested metric: {requested_metric}")

    calculation_keywords = {"calculate", "how much is", "what is", "formula for", "compute"}
    is_calculation_request = any(word in doc.text for word in calculation_keywords)

    if requested_metric and is_calculation_request:
        if requested_metric in financial_formulas:
            formula_info = financial_formulas[requested_metric]
            department = None

            # Try to identify a department or specific context from the question
            for ent in doc.ents:
                # Use a broader range of entity types or custom entity recognition for departments
                # This might need fine-tuning based on your actual department names in data
                if ent.label_ in ["ORG", "NORP", "GPE", "PRODUCT", "FAC"]:
                    normalized_ent = normalize_col(ent.text)
                    # Check against the keys within detailed_summary, especially under 'revenue' or other metrics
                    # to identify actual departments present in the data.
                    possible_departments_in_data = set()
                    for metric_key, metric_data in detailed_summary.items():
                        for k, v in metric_data.items():
                            if isinstance(v, (int, float)) and 'dept' in k:  
                                possible_departments_in_data.add(k)
                            elif isinstance(v, dict):  # For nested departments
                                possible_departments_in_data.add(k)  # Add parent key (e.g., 'sales')
                                for nested_key in v.keys():
                                    if 'dept' in nested_key: 
                                        possible_departments_in_data.add(nested_key)

                    match_list_for_department = get_close_matches(normalized_ent, list(possible_departments_in_data),
                                                                  n=1, cutoff=0.7)
                    if match_list_for_department:
                        department = match_list_for_department[0]
                        logging.info(
                            f"NLP identified potential department for calculation: {department} from entity '{ent.text}'")
                        break

            # If the user explicitly asks for the formula, provide it
            if "formula for" in question.lower() or "how is" in question.lower() and "calculated" in question.lower():
                return f"The formula for {requested_metric.replace('_', ' ').title()} is: {formula_info['formula']}"

            # Otherwise, perform the calculation
            calculated_value = formula_info["calculation_func"](detailed_summary, department)
            if calculated_value is not None:
                if department:
                    return f"The {requested_metric.replace('_', ' ').title()} for {department.replace('_', ' ').title()} is: ${calculated_value:,.2f}"
                else:
                    return f"The total {requested_metric.replace('_', ' ').title()} is: ${calculated_value:,.2f}"
            else:
                return f"Could not calculate {requested_metric.replace('_', ' ').title()} with the available data for the specified context."

    # If not a calculation request or calculation failed, proceed with direct data lookup
    if not requested_metric or requested_metric not in detailed_summary:
        logging.info(f"No valid metric found or metric '{requested_metric}' not in summary for question: {question}")
        return None

    # Collect all possible parent and child keys from the detailed_summary for NLP matching.
    possible_parent_keys = set()
    possible_child_keys = set()
    for metric_data_val in detailed_summary.values():
        for key, value in metric_data_val.items():
            if isinstance(value, dict):
                possible_parent_keys.add(key)
                for child_key in value.keys():
                    possible_child_keys.add(child_key)
            elif key != 'total' and not key.startswith('category_'):
                possible_parent_keys.add(key)

    logging.info(f"Possible parent keys from data: {possible_parent_keys}")
    logging.info(f"Possible child keys from data: {possible_child_keys}")

    parent_match_from_nlp = None
    child_match_from_nlp = None

    for chunk in doc.noun_chunks:
        chunk_text = normalize_col(chunk.text)
        match_p_list = get_close_matches(chunk_text, list(possible_parent_keys), n=1, cutoff=0.75)
        if match_p_list:
            matched_key = match_p_list[0]
            if not parent_match_from_nlp or len(matched_key) > len(parent_match_from_nlp):
                parent_match_from_nlp = matched_key
                logging.info(f"NLP matched parent: '{parent_match_from_nlp}' from chunk: '{chunk.text}'")

        match_c_list = get_close_matches(chunk_text, list(possible_child_keys), n=1, cutoff=0.75)
        if match_c_list:
            matched_key = match_c_list[0]
            if not child_match_from_nlp or len(matched_key) > len(child_match_from_nlp):
                child_match_from_nlp = matched_key
                logging.info(f"NLP matched child: '{child_match_from_nlp}' from chunk: '{chunk.text}'")

    for token in doc:
        token_text = normalize_col(token.text)
        match_p_list = get_close_matches(token_text, list(possible_parent_keys), n=1, cutoff=0.8)
        if match_p_list:
            matched_key = match_p_list[0]
            if not parent_match_from_nlp or len(matched_key) > len(parent_match_from_nlp):
                parent_match_from_nlp = matched_key
                logging.info(f"NLP matched parent (token): '{parent_match_from_nlp}' from token: '{token.text}'")

        match_c_list = get_close_matches(token_text, list(possible_child_keys), n=1, cutoff=0.8)
        if match_c_list:
            matched_key = match_c_list[0]
            if not child_match_from_nlp or len(matched_key) > len(child_match_from_nlp):
                child_match_from_nlp = matched_key
                logging.info(f"NLP matched child (token): '{child_match_from_nlp}' from token: '{token.text}'")

    logging.info(f"Final NLP matches: Parent='{parent_match_from_nlp}', Child='{child_match_from_nlp}'")

    metric_data = detailed_summary[requested_metric]
    logging.info(f"Data for requested metric '{requested_metric}': {metric_data}")

    if parent_match_from_nlp and child_match_from_nlp:
        logging.info(f"Attempting to retrieve with parent='{parent_match_from_nlp}' and child='{child_match_from_nlp}'")
        if parent_match_from_nlp in metric_data and isinstance(metric_data[parent_match_from_nlp], dict) and \
                child_match_from_nlp in metric_data[parent_match_from_nlp]:
            val = metric_data[parent_match_from_nlp][child_match_from_nlp]
            logging.info(f"Found value: ${val:,.2f}")
            return f"${val:,.2f}"
        logging.info("Parent-child path not found directly for nested dictionary.")

    if parent_match_from_nlp:
        logging.info(f"Attempting to retrieve with only parent='{parent_match_from_nlp}'")
        actual_parent_key_in_data = None
        match_list_for_data_key = get_close_matches(parent_match_from_nlp, list(metric_data.keys()), n=1, cutoff=0.6)
        if match_list_for_data_key:
            actual_parent_key_in_data = match_list_for_data_key[0]
            logging.info(
                f"Fuzzy matched NLP parent '{parent_match_from_nlp}' to actual data key '{actual_parent_key_in_data}'")

        if actual_parent_key_in_data and actual_parent_key_in_data in metric_data:
            data = metric_data[actual_parent_key_in_data]
            if isinstance(data, (int, float)):
                logging.info(f"Found direct parent value: ${data:,.2f}")
                return f"${data:,.2f}"
            elif isinstance(data, dict):
                total_for_parent = sum(v for v in data.values() if isinstance(v, (int, float)))
                if total_for_parent > 0:
                    logging.info(f"Summed values for parent '{actual_parent_key_in_data}': ${total_for_parent:,.2f}")
                    return f"${total_for_parent:,.2f}"
        logging.info("Parent match yielded no direct or summable value for metric.")

    if child_match_from_nlp:
        logging.info(f"Attempting to retrieve with only child='{child_match_from_nlp}'")
        if child_match_from_nlp in metric_data and isinstance(metric_data[child_match_from_nlp], (int, float)):
            val = metric_data[child_match_from_nlp]
            logging.info(f"Found direct child value: ${val:,.2f}")
            return f"${val:,.2f}"

        sum_for_child = 0
        found_child_in_parent = False
        for key, value_obj in metric_data.items():
            if isinstance(value_obj, dict) and child_match_from_nlp in value_obj and isinstance(
                    value_obj[child_match_from_nlp], (int, float)):
                sum_for_child += value_obj[child_match_from_nlp]
                found_child_in_parent = True
        if found_child_in_parent:
            logging.info(f"Summed values for child '{child_match_from_nlp}' across parents: ${sum_for_child:,.2f}")
            return f"${sum_for_child:,.2f}"
        logging.info("Child match yielded no direct or summable value for metric.")

    if any(word in doc.text for word in ["total", "overall", "all", "sum", "company"]):
        logging.info("Question contains 'total' or similar keyword. Attempting to find overall total.")
        if 'total' in metric_data and isinstance(metric_data['total'], (int, float)):
            logging.info(f"Found explicit 'total' key: ${metric_data['total']:,.2f}")
            return f"${metric_data['total']:,.2f}"

        total_sum = 0
        for key, value_obj in metric_data.items():
            if isinstance(value_obj, (int, float)) and key != 'total':
                total_sum += value_obj
            elif isinstance(value_obj, dict):
                total_sum += sum(v for v in value_obj.values() if isinstance(v, (int, float)))

        if total_sum > 0:
            logging.info(f"Calculated overall sum for '{requested_metric}': ${total_sum:,.2f}")
            return f"${total_sum:,.2f}"
        logging.info("No overall total found or calculable.")

    logging.info(
        f"No specific answer found for question: '{question}' after querying detailed summary. Falling back to LLM.")
    return None


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
def upload():
    uploaded_file = request.files.get('file')
    question = request.form.get('question')

    if not uploaded_file or not question:
        return "Missing file or question", 400

    try:
        session_id = session.setdefault('session_id', str(uuid4()))
        file_bytes = uploaded_file.read()
        uploaded_files[session_id] = file_bytes  # Store file bytes in-memory

        df = pd.read_excel(io.BytesIO(file_bytes), engine='openpyxl', header=[0, 1])
        df.columns = ['_'.join(filter(None, [str(i) for i in col])) for col in df.columns]
        logging.info(f"Flattened DataFrame columns after upload: {df.columns.tolist()}")
        logging.info(f"DataFrame head:\n{df.head().to_string()}")

        summary = extract_financial_summary(df)
        table_text = dataframe_to_text_table(df)
        error_cells = get_error_cells(file_bytes)
        if error_cells:
            summary += f"\n\nNote: Errors found in cells: {', '.join(error_cells)}"

        # Detailed summaries are generated here and stored in session for the first time
        detailed_summaries = extract_detailed_summaries(df)
        session[
            'detailed_summaries'] = detailed_summaries  # Still store for initial chat turn's direct access optimization
        session['summary'] = summary
        session['table'] = table_text

        # Initialize conversation history, providing the summary and initial question to the LLM
        session['conversation'] = [{"role": "user",
                                    "content": f"Here is a summary of the financial data: {summary}\nNow, please answer this question: {question}"}]

        direct_answer = query_detailed_summary(question, detailed_summaries)

        if direct_answer:
            answer = direct_answer
            logging.info(f"Direct answer found for initial question: {answer}")
        else:
            logging.info("Direct answer not found for initial question, falling back to LLM.")
            prompt = (
                f"You are a financial assistant. Based on the provided financial summary, "
                f"answer the following question. If the answer is a numerical value, "
                f"please provide only the number with a '$' sign and comma formatting. "
                f"Otherwise, provide a concise textual answer, explaining if necessary.\n"
                f"Financial Summary:\n{summary}\n\n"
                f"Question: {question}\nAnswer:"
            )
            result = pipe(prompt, max_new_tokens=100, do_sample=False)
            answer = result[0]['generated_text'].strip()
            logging.info(f"LLM generated answer for initial question: {answer}")

        session['conversation'].append({"role": "assistant", "content": answer})

        return render_template(
            'chat.html',
            answer=answer,
            filename=uploaded_file.filename,
            question=question,
            data_preview=df.head().to_html(index=False),
            table_text=table_text
        )

    except Exception as e:
        logging.exception("File processing error during upload.")
        return f"Error processing file: {str(e)}", 500


@app.route('/chat', methods=['POST'])
def chat():
    question = request.form.get('question')
    if not question:
        return "No question provided", 400

    session_id = session.get('session_id')
    if not session_id or session_id not in uploaded_files:
        logging.warning(f"Session expired or file not found for session_id: {session_id}. Please re-upload.")
        return "Session expired or missing data. Please re-upload the file.", 400

    file_bytes = uploaded_files[session_id]
    try:
        # Re-read DataFrame and re-extract summaries on each chat request for robustness
        df = pd.read_excel(io.BytesIO(file_bytes), engine='openpyxl', header=[0, 1])

        # Flatten MultiIndex columns to single strings
        df.columns = ['_'.join(filter(None, [str(i) for i in col])) for col in df.columns]

        # Ensure column names are unique to avoid parsing issues
        df.columns = pd.io.parsers.ParserBase({'names': df.columns})._maybe_dedup_names(df.columns)

        detailed_summaries = extract_detailed_summaries(df)
        summary = extract_financial_summary(df)  # Regenerate summary if needed for LLM
        table_text = dataframe_to_text_table(df)  # Regenerate table text if needed for template

        conversation = session.get('conversation', [])

        direct_answer = query_detailed_summary(question, detailed_summaries)

        if direct_answer:
            answer = direct_answer
            logging.info(f"Direct answer found in chat: {answer}")
        else:
            conversation.append({"role": "user", "content": question})
            turns = "\n".join(f"{msg['role'].capitalize()}: {msg['content']}" for msg in conversation)

            prompt = (
                f"You are a financial assistant. Based on the provided financial summary and conversation history, "
                f"answer the following question. If the answer is a numerical value, "
                f"please provide only the number with a '$' sign and comma formatting. "
                f"Otherwise, provide a concise textual answer, explaining if necessary.\n\n"
                f"Financial Summary:\n{summary}\n\n"  # Use the regenerated summary
                f"Conversation History:\n{turns}\n\n"
                f"Answer:"
            )
            result = pipe(prompt, max_new_tokens=100, do_sample=False)
            answer = result[0]['generated_text'].strip()
            logging.info(f"LLM generated answer in chat: {answer}")

            conversation.append({"role": "assistant", "content": answer})

        session['conversation'] = conversation
        session['summary'] = summary  # Update summary in session
        session['detailed_summaries'] = detailed_summaries  # Update detailed_summaries in session

        return render_template(
            'chat.html',
            answer=answer,
            question=question,
            table_text=table_text  # Use the regenerated table text
        )

    except Exception as e:
        logging.exception("Chat processing error.")
        return f"Error during chat processing: {str(e)}", 500


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
